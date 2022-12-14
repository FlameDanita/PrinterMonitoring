import os
import subprocess
import pandas as pd #pip install pandas
import xlwt #pip install xlwt
from openpyxl import load_workbook #pip install openpyxldd
from tqdm import tqdm
import datetime

def get_IOD(wb, ser):
    IOD = {}
    work_sheet_IOD = wb['IOD']
    flag = False
    counter = 0
    for rows in work_sheet_IOD.iter_rows(values_only=True):
        if rows[2] == ser[:2] and flag == False:
            flag = True
        elif flag == True and counter < 9:
            if rows[0] and rows[0] != 'Модель':
                IOD[rows[0]] = rows[1]
            counter += 1
    return IOD

def search_cell(name_cell, work_sheet):
    for cell in range(1, 100):
        if work_sheet.cell(row = 1, column = cell).value == name_cell:
            break
    print('Столбец "', name_cell, '": ', cell, sep = '')
    return cell

if __name__ == "__main__":
    wb = load_workbook('./printers.xlsx')
    print(wb.sheetnames)
    work_sheet = wb['MAIN']
    cell_ser = search_cell('Серийный номер', work_sheet) - 1
    cell_IP = search_cell('IP', work_sheet) - 1

    start_row = 2
    current_sheet = 3
    for row in tqdm(work_sheet.iter_rows(min_row=2, values_only=True)):
        work_sheet_pr = wb[wb.sheetnames[current_sheet]]
        if row[0]:
            pass
        else:
            break
        result = []
        IOD = {} 
        try:
            IOD = get_IOD(wb, ser=row[cell_ser])
            for key, value in IOD.items():
                res = subprocess.check_output("./SnmpGet.exe -r:" + str(row[cell_IP]) + 
                                        ' -o:' + value).decode("cp866").split('\n')
                result.append(res[-2].split('=')[-1])
                IOD[key] = str(res[-2].split('=')[-1]).strip('\r')
            cur_row = 1
            flag = False
            for row_pr in work_sheet_pr.iter_rows(values_only=True):
                if row_pr[0] in IOD:
                    try:
                        work_sheet_pr.cell(row=cur_row, column=2).value = int(IOD[row_pr[0]])
                    except:
                        work_sheet_pr.cell(row=cur_row, column=2).value = IOD[row_pr[0]]
                if row_pr[0] == 'Status':
                    work_sheet_pr.cell(row=cur_row, column=2).value = 'online'
                if flag:
                    try:
                        tmp = str(row_pr[0].split()[0])
                    except:
                        tmp = row_pr[0]
                    if tmp == str(datetime.datetime.now().strftime("%d.%m.%Y")) or not tmp:
                        work_sheet_pr.cell(row=cur_row, column=1).value = str(datetime.datetime.now().strftime("%d.%m.%Y"))
                        work_sheet_pr.cell(row=cur_row, column=2).value = int(IOD['Оставшееся число копий тонера'])/int(IOD['Максимальное число копий тонера'])
                        work_sheet_pr.cell(row=cur_row, column=3).value = int(IOD['Кол-во напечатанных страниц'])
                        flag = False
                if row_pr[0] == 'Дата':
                    flag = True
                cur_row += 1
        except:
            work_sheet_pr = wb[wb.sheetnames[current_sheet]]
            cur_row = 1
            for row_pr in work_sheet_pr.iter_rows(values_only=True):
                if row_pr[0] == 'Status':
                    if IOD['Серийный номер принтера'][0] == '%':
                        work_sheet_pr.cell(row=cur_row, column=2).value = 'offline'
                        print(IOD)
                        print("Printer", row[1], 'offline')
                    else:
                        work_sheet_pr.cell(row=cur_row, column=2).value = 'failed'
                        print(IOD)
                        print("Printer", row[1], 'failed')
                cur_row += 1 
        current_sheet += 1

    wb.save('./printers.xlsx')